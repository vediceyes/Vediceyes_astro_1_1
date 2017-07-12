import sys
import time
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import swisseph as swe
import collections
from Swiss_eph_constants import *
from geopy.geocoders import Nominatim
from ChartCalculations import Chart
import pytz
from pytz import timezone
import datetime
import openpyxl
import glob
from openpyxl import Workbook


birth_chart_list = glob.glob("/Users/anandramamurthy/PycharmProjects/Vediceyes_astro_1_1/Birthdata*.xlsx")

print(birth_chart_list)

from openpyxl import load_workbook


Timezone_list=[]

Timezone_list= pytz.all_timezones

import ui_samainwindow
import sys


class UI(QMainWindow, ui_samainwindow.Ui_SAMainWindow,Chart,Nominatim):
    def __init__(self, parent=None):
        super(UI, self).__init__(parent)
        self.__index = 0
        self.setupUi(self)

        #Class creation
        self.geolocator = Nominatim(scheme='http')
        self.Hscope = Chart()

        # Object connections:
        self.Timezone.addItems(Timezone_list)
        self.Chartname.addItems(birth_chart_list)
        self.savebirth.clicked.connect(self.Create_chartdata)
        self.LoadChart.clicked.connect(self.Load_Chart)

    def Create_chartdata(self):

        Chartbook = Workbook()
        Chartsheet = Chartbook.active

        Chartsheet.append(["Name", "DOB Month", "DOB Date", "DOB Year", "Time(Hour)", "Minutes","Location","Time Zone", "Longitude" "Lattitude"])

        Name = self.NameInput.text()

        year = self.DateInput.date().year()
        month = self.DateInput.date().month()
        day = self.DateInput.date().day()

        hour = self.TimeInput.time().hour()
        minute = self.TimeInput.time().minute()

        Location = self.locstring.text()
        geolocator = Nominatim(scheme='http')
        Find_location = geolocator.geocode(Location)
        tzvalue = self.Timezone.currentText()
        Long = Find_location.longitude
        lat = Find_location.latitude

        Chartsheet['D2'] = year
        Chartsheet['B2'] = month
        Chartsheet['C2'] = day
        Chartsheet['A2'] = Name

        Chartsheet['E2'] = hour
        Chartsheet['F2'] = minute

        Chartsheet['G2'] = Location
        Chartsheet['H2'] = tzvalue
        Chartsheet['I2'] = Long
        Chartsheet['J2'] = lat


        self.latDisplay.setText(str(Find_location.latitude))
        self.longDisplay.setText(str(Find_location.longitude))


        Chartbook.save("Birthdata_"+Name+".xlsx")


    def CalculateBirth(self,year,month,day, name, hour, minute, Location, tzvalue, Long, lat):

        wb = load_workbook('BirthChart.xlsx')
        ws = wb.active

        for row in ws['A2:N9']:
            for cell in row:
                cell.value = None

        for row in ws['B16:M16']:
            for cell in row:
                cell.value = None

        for row in ws['B26:M26']:
            for cell in row:
                cell.value = None


        self.Hscope.Setup_eph(Long, lat)

        local_tz = timezone(tzvalue)

        t = local_tz.localize(datetime.datetime(year, month, day, hour, minute))

        t_utc = t.astimezone(pytz.UTC)

        hour_utc= t_utc.hour
        minute_utc= t_utc.minute

        Bday_accurate = self.Hscope.DateTime(year, month, day, hour_utc, minute_utc, 0)

        print(Bday_accurate)

        (HouseDict, HousePosList) = self.Hscope.CaclHouses(Bday_accurate, Long, lat)

        self.Hscope.CalcPlanets(HouseDict, Bday_accurate, ws)

        self.Hscope.getASC(HouseDict)

        self.title.setText(name)

        z = 0
        for i in SH_List:
            self.HouseTable.setItem(0, z, QTableWidgetItem(ws[i + str(26)].value))
            z = z + 1

        z = 0
        for i in SH_List:
            self.Sign_table.setItem(0, z, QTableWidgetItem(ws[i + str(16)].value))
            z = z + 1

        for y in range (2,10):
            z = 0
            for i in Dignity_List:
                self.Planet_Dignity.setItem(y-2, z, QTableWidgetItem(ws[i + str(y)].value))
                z = z + 1

        wb.save('BirthChart.xlsx')

    def Load_Chart(self):

        chart= load_workbook(self.Chartname.currentText())
        chart_ws= chart.active

        year = chart_ws['D2'].value
        month = chart_ws['B2'].value
        day = chart_ws['C2'].value
        name = chart_ws['A2'].value

        hour = chart_ws['E2'].value
        minute = chart_ws['F2'].value

        Location = chart_ws['G2'].value
        tzvalue = chart_ws['H2'].value
        Long = chart_ws['I2'].value
        lat = chart_ws['J2'].value

        self.CalculateBirth(year,month,day, name, hour, minute, Location, tzvalue, Long, lat)

        # upload to table



app = QApplication(sys.argv)

window = UI()
window.show()
app.exec_()

