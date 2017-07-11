import openpyxl

from openpyxl import load_workbook


#wb = load_workbook('GeoLiteCity-Location.xlsx')

#ws = wb.active

#d = ws['c123454'].value

#print(d)


from geopy.geocoders import Nominatim
geolocator = Nominatim(scheme='http')
location = geolocator.geocode("guntur, andra pradesh, India")

print(location.latitude, location.longitude)